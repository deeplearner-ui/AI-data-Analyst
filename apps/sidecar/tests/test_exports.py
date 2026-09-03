import base64
import json
from io import BytesIO
from pathlib import Path
from zipfile import ZipFile

import pandas as pd
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from aida_sidecar.api import app


client = TestClient(app)
HEADERS = {"Authorization": "Bearer development-token"}


def create_export_project(tmp_path: Path) -> tuple[Path, str]:
    project = tmp_path / "export-project"
    source = tmp_path / "export.csv"
    pd.DataFrame({"名称": ["甲", "乙"], "值": [10, 20], "文本": ["=1+1", "安全"]}).to_csv(source, index=False)
    assert client.post("/api/projects", headers=HEADERS, json={"directory": str(project), "name": "导出项目", "language": "zh-CN"}).status_code == 200
    imported = client.post("/api/datasets/import", headers=HEADERS, json={"projectDirectory": str(project), "path": str(source)}).json()
    return project, imported["version"]["id"]


def test_dataset_report_and_reproducibility_exports(tmp_path: Path) -> None:
    project, version_id = create_export_project(tmp_path)

    csv_response = client.post("/api/datasets/export", headers=HEADERS, json={"projectDirectory": str(project), "versionId": version_id, "format": "csv"})
    assert csv_response.status_code == 200
    csv_bytes = base64.b64decode(csv_response.json()["contentBase64"])
    assert csv_bytes.startswith(b"\xef\xbb\xbf")
    assert "'=1+1" in csv_bytes.decode("utf-8-sig")

    xlsx_response = client.post("/api/datasets/export", headers=HEADERS, json={"projectDirectory": str(project), "versionId": version_id, "format": "xlsx"})
    workbook = load_workbook(BytesIO(base64.b64decode(xlsx_response.json()["contentBase64"])), read_only=True, data_only=False)
    assert workbook.sheetnames == ["Data", "AIDA Metadata"]
    assert workbook["Data"]["C2"].value == "'=1+1"

    sections = [{"id": "summary", "title": "分析摘要", "markdown": "## 结论\n\n- 数据质量良好\n- **均值**可供参考", "resultIds": ["stat_1"], "chartIds": ["chart_1"]}]
    report_payload = {"projectDirectory": str(project), "title": "导出/报告", "sections": sections, "language": "zh-CN", "template": "management", "versionId": version_id, "planId": None}
    html_response = client.post("/api/reports/export", headers=HEADERS, json={**report_payload, "format": "html"})
    assert html_response.status_code == 200
    assert html_response.json()["filename"] == "导出_报告.html"
    html = base64.b64decode(html_response.json()["contentBase64"]).decode("utf-8")
    assert version_id in html and "stat_1" in html and "chart_1" in html
    assert "管理层摘要" in html and "REPORT V3" in html

    pdf_response = client.post("/api/reports/export", headers=HEADERS, json={**report_payload, "format": "pdf"})
    assert pdf_response.status_code == 200
    assert base64.b64decode(pdf_response.json()["contentBase64"]).startswith(b"%PDF-")

    plan = client.post("/api/ai/plan", headers=HEADERS, json={"projectDirectory": str(project), "versionId": version_id, "goal": "验证引用", "includeSamples": False, "language": "zh-CN"}).json()["plan"]
    invalid_reference = client.post("/api/reports/export", headers=HEADERS, json={**report_payload, "format": "html", "planId": plan["id"]})
    assert invalid_reference.status_code == 400
    assert "stat_1" in invalid_reference.json()["detail"]

    bundle_response = client.post("/api/reports/reproducibility", headers=HEADERS, json={**report_payload, "includeData": False, "dataFormat": "csv"})
    assert bundle_response.status_code == 200
    with ZipFile(BytesIO(base64.b64decode(bundle_response.json()["contentBase64"]))) as bundle:
        names = set(bundle.namelist())
        assert {"README.txt", "manifest.json", "versions.json", "analysis/plan.json", "analysis/artifacts.json", "environment.json", "report/report.html", "report/report.pdf"} <= names
        assert not any(name.startswith("data/") for name in names)
        versions = json.loads(bundle.read("versions.json"))
        assert all("storagePath" not in version for version in versions)
        assert str(project).encode() not in bundle.read("versions.json")

    data_bundle = client.post("/api/reports/reproducibility", headers=HEADERS, json={**report_payload, "includeData": True, "dataFormat": "xlsx"}).json()
    with ZipFile(BytesIO(base64.b64decode(data_bundle["contentBase64"]))) as bundle:
        assert "data/current-version.xlsx" in bundle.namelist()
